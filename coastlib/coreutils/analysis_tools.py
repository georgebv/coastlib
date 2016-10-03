import pandas as pd
import scipy.io
import datetime
import detect_peaks
import windrose
import matplotlib.cm as cm
import matplotlib.pyplot as plt
import numpy as np
import statsmodels.api as sm
import math
import openpyxl


def joint_probability(df, **kwargs):
    """
    Generates a joint probability table of 2 variables.

    Parameters
    ----------
    df:
    kwargs:
    """
    val1 = kwargs.get('val1', 'Hs')
    val2 = kwargs.get('val2', 'Tp')
    binsize1 = kwargs.get('binsize1', 0.3)
    binsize2 = kwargs.get('binsize2', 4)
    savepath = kwargs.get('savepath', None)
    savename = kwargs.get('savename', 'Joint Probability')

    a = df[pd.notnull(df[val1])]
    a = a[pd.notnull(a[val2])]
    vals1 = a[val1]
    vals2 = a[val2]
    bins1 = math.ceil(vals1.max() / binsize1)
    bins2 = math.ceil(vals2.max() / binsize2)
    jp_table = np.zeros((bins1, bins2))
    tot = len(vals1)
    for i in range(0, bins1 - 1):
        bin1_low = i * binsize1
        bin1_up = bin1_low + binsize1
        for j in range(0, bins2 - 1):
            bin2_low = j * binsize2
            bin2_up = bin2_low + binsize2
            p1 = (((vals1 >= bin1_low) & (vals1 < bin1_up)).sum()) / tot
            p2 = (((vals2 >= bin2_low) & (vals2 < bin2_up)).sum()) / tot
            jp_table[i, j] = p1 * p2
    if savepath is not None:
        book = openpyxl.Workbook()
        ws1 = book.active
        ws1.title = val1 + ' vs ' + val2
        for i in range(0, bins1 - 1):
            bin1_low = i * binsize1
            bin1_up = bin1_low + binsize1
            ws1.cell(column=1, row=i+2, value=str(bin1_low) + '-' + str(bin1_up))
            for j in range(0, bins2 - 1):
                bin2_low = j * binsize2
                bin2_up = bin2_low + binsize2
                ws1.cell(column=j+2, row=1, value=str(bin2_low) + '-' + str(bin2_up))
                ws1.cell(column=j+2, row=i+2, value=jp_table[i, j])
        book.save(filename=savepath + '\\' + savename + '.xlsx')
    else:
        return jp_table
